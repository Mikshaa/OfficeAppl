import os
import time
import rec_rc
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import * 
from PyQt5.QtCore import *
import urllib
import xlwings
import pandas as pe
import openpyxl
import datetime
import xlrd
from math import floor


curMode = 'Device'
curSecondMode = 'Consolid'
curDeviceCode = ''
curAmount = 0
curFile = ''



inputFilesPath = ''
outputFilesPath = ''
contractFilePath = ''
curContractFile = ''
contractDevices = []

devicesList = []


var_1 = ''
var_2 = ''
var_3 = ''
var_4 = ''
var_5 = ''
var_6 = ''

checkForFormula = ['A', 'I', 'i', 'U', 'E', 'X', 'K', 'L', 'O', 'S', 'G','У','.','W','З','В','М','О','Е','Р','С','А','Л','К','Т','Ь','И']
checkDeviceCode = True
zeroError = False
contractFilesError = False

def connect(mode = 'check'):
    if mode == 'check':
        host = 'http://google.com'
    elif mode == 'recheck':
        host = 'https://ya.ru/'
    try:
        urllib.request.urlopen(host)
        connected = True
    except:
        connected = False
    if not connected:
        ui.showErrorMessagebox(mode='connection')
def setVars(device_code):
    received_variables = device_code.split('-')
    global var_1, var_2, var_3, var_4, var_5, var_6, checkDeviceCode
    checkDeviceCode = True
    if received_variables[2] == 'Р':
        var_1 = '"Р"'
    elif received_variables[2] == 'Э':
        var_1 = '"Э"'
    else:
        checkDeviceCode = False
        ui.showErrorMessagebox(text='Неправильно указан\nтип привода')
        return

    if received_variables[3] == 'П':
        var_2 = '"П"'
    elif received_variables[3] == 'Л':
        var_2 = '"Л"'
    else:
        checkDeviceCode = False
        ui.showErrorMessagebox(text='Неправильно указана\nсторона исполнения')
        return

    if received_variables[4] == 'Т1':
        var_3 = '"Т1"'
    elif received_variables[4] == 'Т2':
        var_3 = '"Т2"'
    else:
        checkDeviceCode = False
        ui.showErrorMessagebox(text='Неправильно указан\nтип ткани')
        return

    if received_variables[5][0]=='Ш':
        var_4 = received_variables[5][1:]
        if int(var_4)>0 and int(var_4)<=240:
            pass
        else:
            checkDeviceCode = False
            ui.showErrorMessagebox(text='Неправильно указана\nширина')
            return
    else:
        checkDeviceCode = False
        ui.showErrorMessagebox(text='Неправильно указана\nширина1')
        return


    if received_variables[6][0]=='В':
        var_5 = received_variables[6][1:]
        if int(var_5)>0 and int(var_5)<=500:
            pass
        else:
            checkDeviceCode = False
            ui.showErrorMessagebox(text='Неправильно указана\nвысота')
            return
    else:
        checkDeviceCode = False
        ui.showErrorMessagebox(text='Неправильно указана\nвысота')
        return

    if received_variables[7] == 'КР1':
        var_6 = '"КР1"'
    elif received_variables[7]=='КР2':
        var_6 = '"КР2"'
    else:
        checkDeviceCode = False
        ui.showErrorMessagebox(text='Неправильно указан\nтип крепления')
        return

def correctFinalFile(filePath, length):
    row = 1
    wb = app.books.open(filePath)
    ws = wb.sheets[0]
    while row < length:
        if ws.range(f'b{row}').value == 0:
            ws.range(f'a{row}').delete(shift='up')
            ws.range(f'b{row}').delete(shift='up')
        else:
            if isinstance(ws.range(f'b{row}').value, float):
                ws.range(f'b{row}').value = floor(float(ws.range(f'b{row}').value*1000))/1000
            ws.range(f'b{row}').value = ws.range(f'b{row}').value
            row += 1
    wb.save()
    wb.close()
def generateDeviceOutuput(mode='device', contractAmount = None):
    global curDeviceCode
    global devicesList
    global checkDeviceCode
    global curAmount
    global contractDevices
    if mode == 'device':
        curAmount = ui.lineEditAmount.text()
        ui.deviceCodeChanged(ui.comboBoxGetDevice.currentText(), mode='device')
    elif mode == 'contract':
        curAmount = contractAmount

    if not checkDeviceCode:
        return
    checkZeros(mode='device')
    if zeroError:
        return
    if inputFilesPath == '':
        ui.showErrorMessagebox(text='Отсутствует директория\nс файлами-шаблонов')
        return
    if outputFilesPath == '':
        ui.showErrorMessagebox(text='Отсутствует директория\nсохранения файлов')
        return
    if int(curAmount) < 1:
        ui.showErrorMessagebox(text='Количесто устройств\n<1')
        return
    if curDeviceCode == '':
        return
    peData = pe.DataFrame(pe.read_excel(f'{inputFilesPath}/{curFile}'),columns=['Наименование ВП', 'Количество', 'Примечание'])
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in range(len(peData)):
        if str(peData['Примечание'][row]) == 'nan':
            ws[f'A{row + 1}'] = str(peData['Наименование ВП'][row])
            ws[f'B{row + 1}'] = float(peData['Количество'][row]) * int(curAmount)
        else:
            if str(peData['Количество'][row]) != 'nan':
                ws[f'A{row + 1}'] = str(peData['Наименование ВП'][row])
                ws[f'B{row + 1}'] = float(peData['Количество'][row]) * int(curAmount)
            else:
                ws[f'A{row + 1}'] = str(peData['Наименование ВП'][row])
                ws[f'B{row + 1}'] = pasteVarsInFormula(str(peData['Примечание'][row]))
    finalSavePath = f'{outputFilesPath}/{curDeviceCode}_{datetime.datetime.now().strftime("%Y-%m-%d")}_{curAmount}.xlsx'
    wb.save(finalSavePath)
    if mode=='contract':
        contractDevices.append([finalSavePath, len(peData)+1])



    if mode == 'device':
        row = 1
        print(finalSavePath)
        app = xlwings.App(visible=False, add_book=False)
        wb = app.books.open(finalSavePath)
        ws = wb.sheets[0]
        while row < len(peData)+1:
            if ws.range(f'b{row}').value == 0:
                ws.range(f'a{row}').delete(shift='up')
                ws.range(f'b{row}').delete(shift='up')
            else:
                if isinstance(ws.range(f'b{row}').value, float):
                    ws.range(f'b{row}').value = floor(float(ws.range(f'b{row}').value*1000))/1000
                else:
                    ws.range(f'b{row}').value = ws.range(f'b{row}').value
                row += 1
        wb.save()
        wb.close()
        app.quit()
    elif mode == 'contract':
        pass
    if mode=='device':
        ui.showFinalMessage()

def genarateContractOutput():
    global curContractFile, contractFilePath, curSecondMode, curDeviceCode, zeroError, contractFilesError, contractDevices, app
    filesError = False
    contractDevices = []
    contractData = pe.DataFrame(pe.read_excel(contractFilePath,header=None))
    app = xlwings.App(visible=False)
    wb = app.books.open(f'{contractFilePath}')
    ws = wb.sheets[0]
    for row in range(len(contractData)):
        curDeviceCode = contractData[0][row]
        ui.deviceCodeChanged(curDeviceCode, mode='contract')
        if contractFilesError:
            ws.range(f'a{row + 1}:b{row + 1}').color = (201, 40, 40)
            filesError = True
    wb.save()
    wb.close()
    app.quit()
    if filesError:
        ui.showErrorMessagebox(text='Файл-шаблон\nотсутствует')
        app = xlwings.App(visible=True, add_book=False)
        wb = app.books.open(f'{contractFilePath}')
        return

    checkZeros(mode='contract')
    if zeroError:
        return
    for row in range(len(contractData)):
        deviceCode = contractData[0][row]
        contractAmount = contractData[1][row]
        ui.deviceCodeChanged(deviceCode)
        generateDeviceOutuput(mode='contract', contractAmount=contractAmount)
    time.sleep(2)
    app = xlwings.App(visible=False, add_book=False)
    for device in contractDevices:
        correctFinalFile(device[0], device[1])
    app.quit()
    if curSecondMode == 'Separate':
        ui.showFinalMessage(text='Файлы сгенерированы')
    elif curSecondMode == 'Consolid':
        wbDataFrames = []
        for device in contractDevices:
            wbData = pe.DataFrame(pe.read_excel(device[0],header=None))
            wbDataFrames.append(wbData)
            os.remove(device[0])
        wbDataFinal = wbDataFrames[0]
        wbDataFinalList = wbDataFinal[0].tolist()
        wbValues = wbDataFinal[1].tolist()
        del wbDataFrames[0]
        for i in wbDataFrames:
            for j in range(len(i)):
                if i[0][j] in wbDataFinalList:
                    wbValues[wbDataFinalList.index(i[0][j])] = float(i[1][j])+float(wbValues[wbDataFinalList.index(i[0][j])])
                else:
                    wbDataFinalList.append(str(i[0][j]))
                    wbValues.append(float(i[1][j]))
        wb_final = openpyxl.Workbook()
        ws_final = wb_final.active
        for i in range(len(wbDataFinalList)):
            ws_final[f'A{i+1}'] = wbDataFinalList[i]
            ws_final[f'B{i+1}'] = wbValues[i]
        wb_final.save(f'{outputFilesPath}/{contractFilePath[contractFilePath.rfind("/")+1:contractFilePath.rfind(".")]}.xlsx')
        ui.showFinalMessage(text='Файл договора\nсгенерирован')

def pasteVarsInFormula(formula):
    nInFormula = False
    formula = formula.replace('П1', var_1)
    formula = formula.replace('П2', var_2)
    formula = formula.replace('П3', var_3)
    for i in range(len(formula)):
        if formula[i] == 'Ш' and formula[i-1] not in checkForFormula:
            formula = formula.replace('Ш', var_4)
    for i in range(len(formula)):
        if formula[i] == 'В' and formula[i-1] not in checkForFormula:
            formula = formula.replace('В', var_5)
    formula = formula.replace('П6', var_6)
    print(formula)
    for i in range(len(formula)):
        if formula[i] == 'N' and formula[i-1] not in checkForFormula:
            formula = formula[:i]+str(curAmount)+formula[i+1:]
            nInFormula = True
    print(formula)
    formula = formula.replace(';', ',')
    if nInFormula:
        formula = f'=({formula})'
    else:
        formula = f'=({formula})*{curAmount}'
    return formula

def checkZeros(mode = 'device'):
    global curMode
    global zeroError
    global contractFilePath
    if mode == 'device':
        zeroError = False
        rowsNum = []
        peDatas = pe.DataFrame(pe.read_excel(f'{inputFilesPath}/{curFile}'),columns=['Количество'])
        for row in range(len(peDatas['Количество'])):
            if peDatas['Количество'][row] == 0.0 or isinstance(peDatas['Количество'][row], str):
                zeroError = True
                rowsNum.append(row+2)
        if zeroError:
            ui.showErrorMessagebox(text='Некорректное значение\nв файле шаблона')
            app = xlwings.App(visible=True, add_book=False)
            wb = app.books.open(f'{inputFilesPath}/{curFile}')
            ws = wb.sheets[0]
            for row in rowsNum:
                ws.range(f'a{row}:k{row}').color = (201, 40, 40)
            wb.save()

    elif mode == 'contract':
        zeroError = False
        rowsNum = []
        peDatas = pe.DataFrame(pe.read_excel(f'{contractFilePath}',header=None))
        for row in range(len(peDatas[1])):
            if peDatas[1][row]==0.0 or isinstance(peDatas[1][row], str):
                zeroError = True
                rowsNum.append(row+1)
        if zeroError:
            ui.showErrorMessagebox(text='Некорректное значение\nв файле договора')
            app = xlwings.App(visible=True, add_book=False)
            wb = app.books.open(f'{contractFilePath}')
            ws = wb.sheets[0]
            for row in rowsNum:
                ws.range(f'a{row}:k{row}').color = (201, 40, 40)
            wb.save()

class Ui_MainWindow(object):

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setFixedSize(1200, 400)
        MainWindow.setStyleSheet("background-color: #F5F5F5")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        #############################################

        self.radioButtonDevice = QtWidgets.QRadioButton(self.centralwidget)
        self.radioButtonDevice.setGeometry(QtCore.QRect(40, 120, 130, 40))
        self.radioButtonDevice.setObjectName("radioButton")

        #############################################

        self.radioButtonContract = QtWidgets.QRadioButton(self.centralwidget)
        self.radioButtonContract.setGeometry(QtCore.QRect(40, 160, 130, 40))
        self.radioButtonContract.setObjectName("radioButton_2")

        #############################################

        self.radioButtonConsolid = QtWidgets.QRadioButton(self.centralwidget)
        self.radioButtonConsolid.setGeometry(QtCore.QRect(60, 200, 190, 40))
        self.radioButtonConsolid.setObjectName("radioButton_3")
        self.radioButtonConsolid.setChecked(True)
        self.radioButtonConsolid.setEnabled(False)

        ##############################################

        self.radioButtonDelen = QtWidgets.QRadioButton(self.centralwidget)
        self.radioButtonDelen.setGeometry(QtCore.QRect(60, 240, 130, 40))
        self.radioButtonDelen.setObjectName("radioButton_4")
        self.radioButtonDelen.setEnabled(False)

        #############################################

        self.btngroup1 = QtWidgets.QButtonGroup()
        self.btngroup2 = QtWidgets.QButtonGroup()
        self.btngroup1.addButton(self.radioButtonDevice)
        self.btngroup1.addButton(self.radioButtonContract)
        self.btngroup2.addButton(self.radioButtonConsolid)
        self.btngroup2.addButton(self.radioButtonDelen)
        self.radioButtonDevice.setChecked(True)

        #############################################
        
        self.buttonGenerate = QtWidgets.QPushButton(self.centralwidget)
        self.buttonGenerate.setGeometry(QtCore.QRect(660, 300, 180, 35))
        self.buttonGenerate.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.buttonGenerate.setFocusPolicy(QtCore.Qt.ClickFocus)
        self.buttonGenerate.setStyleSheet("QPushButton{\n"
                                        "border: 2px solid;\n"
                                        "border-color: #B7B7B7;\n"
                                        "border-radius: 6px;\n"
                                        "color: #B7B7B7;\n"
                                        "\n"
                                        "}\n"
                                        "\n"
                                        "")
        self.buttonGenerate.setObjectName("pushButton_3")
        self.buttonGenerate.setEnabled(False)

        #############################################

        self.buttonGenerate_1 = QtWidgets.QPushButton(self.centralwidget)
        self.buttonGenerate_1.setGeometry(QtCore.QRect(660, 215, 180, 35))
        self.buttonGenerate_1.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.buttonGenerate_1.setFocusPolicy(QtCore.Qt.ClickFocus)
        self.buttonGenerate_1.setStyleSheet("QPushButton{\n"
                                          "border: 1px solid;\n"
                                          "border-color: #000000;\n"
                                          "border-radius: 6px;\n"
                                          "color: #000;\n"
                                          "\n"
                                          "}\n"
                                          "\n"
                                          "")
        self.buttonGenerate_1.setObjectName("pushButton_9")

        #############################################

        self.labelGetInput = QtWidgets.QLabel(self.centralwidget)
        self.labelGetInput.setGeometry(QtCore.QRect(290, 80, 340, 30))
        self.labelGetInput.setObjectName("label")
        self.labelGetInput.setStyleSheet("font-size: 21px;\n")

        #############################################
        
        self.buttonGetOutput = QtWidgets.QPushButton(self.centralwidget)
        self.buttonGetOutput.setGeometry(QtCore.QRect(570, 150, 40, 40))
        self.buttonGetOutput.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.buttonGetOutput.setStyleSheet("border: 2px solid;\n"
                                        "border-radius: 4px;\n"
                                        "border-color: #858585;\n"
                                        "background-color: #D9D9D9")
        self.buttonGetOutput.setObjectName("pushButton_4")

        #############################################
        
        self.buttonGetInput = QtWidgets.QPushButton(self.centralwidget)
        self.buttonGetInput.setGeometry(QtCore.QRect(570, 75, 40, 40))
        self.buttonGetInput.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.buttonGetInput.setStyleSheet("border: 2px solid;\n"
                                        "border-radius: 4px;\n"
                                        "border-color: #858585;\n"
                                        "background-color: #D9D9D9")
        self.buttonGetInput.setObjectName("pushButton_5")

        #############################################
        
        self.labelGetOutput = QtWidgets.QLabel(self.centralwidget)
        self.labelGetOutput.setGeometry(QtCore.QRect(415, 155, 140, 30))
        self.labelGetOutput.setObjectName("label_2")
        self.labelGetOutput.setStyleSheet("font-size: 21px;\n")

        #############################################
        
        self.labelDeviceCode = QtWidgets.QLabel(self.centralwidget)
        self.labelDeviceCode.setGeometry(QtCore.QRect(660, 25, 210, 30))
        self.labelDeviceCode.setObjectName("label_3")
        self.labelDeviceCode.setStyleSheet("font-size: 21px;\n")

        #############################################

        self.labelContract = QtWidgets.QLabel(self.centralwidget)
        self.labelContract.setGeometry(QtCore.QRect(375, 300, 210, 30))
        self.labelContract.setObjectName("label_6")
        self.labelContract.setStyleSheet("font-size: 21px;\n")
        self.labelContract.setEnabled(False)

        #############################################

        self.buttonGetContract = QtWidgets.QPushButton(self.centralwidget)
        self.buttonGetContract.setGeometry(QtCore.QRect(570, 298, 40, 40))
        self.buttonGetContract.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.buttonGetContract.setStyleSheet("border: 2px solid;\n"
                                           "border-radius: 4px;\n"
                                            "color: #B7B7B7;\n"
                                           "border-color: #B7B7B7;\n"
                                           "background-color: #F4F4F4")
        self.buttonGetContract.setObjectName("pushButton_4")
        self.buttonGetContract.setEnabled(False)

        #############################################

        self.comboBoxGetDevice = QtWidgets.QComboBox(self.centralwidget)
        self.comboBoxGetDevice.setGeometry(QtCore.QRect(660, 75, 500, 40))
        self.comboBoxGetDevice.setStyleSheet("QComboBox{\n"
                                    "background-color: #fff;\n"
                                    "border: 1px solid;\n"
                                    "border-radius: 4px;\n"
                                    "\n"
                                    "}\n"
                                    "QComboBox::drop-down{\n"
                                    "width: 34px;\n"
                                    "height: 36px;\n"
                                    "top: 0px;\n"
                                    "border: 1px solid;\n"
                                    "border-left: 2px solid;\n"
                                    "border-radius: 3px;\n"
                                    "background-color: #fff;\n"
                                    "\n"
                                    "}\n"
                                    "\n"
                                    "QComboBox::down-arrow{\n"
                                    "image: url(:/img/Polygon 2.svg)\n"
                                    "}\n"
                                    "")
        self.comboBoxGetDevice.setCurrentText("Выберите")
        self.comboBoxGetDevice.setObjectName("comboBox")
        self.comboBoxGetDevice.setEditable(True)

        #############################################
        
        self.lineEditAmount = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEditAmount.setGeometry(QtCore.QRect(660, 150, 50, 40))
        self.lineEditAmount.setValidator(QIntValidator())
        self.lineEditAmount.setStyleSheet("padding-left: 2px;\n"
                                    "top: 0px;\n"
                                    "border: 2px solid;\n"
                                    "border-radius: 3px;\n"
                                    "font-weight: 400;\n"
                                    "background-color: #fff")
        self.lineEditAmount.setObjectName("lineEdit")
        MainWindow.setCentralWidget(self.centralwidget)

        #############################################

        self.line = QtWidgets.QFrame(self.centralwidget)
        self.line.setGeometry(QtCore.QRect(250, 75, 20, 250))
        self.line.setFrameShape(QtWidgets.QFrame.VLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.line_2 = QtWidgets.QFrame(self.centralwidget)
        self.line_2.setGeometry(QtCore.QRect(660, 270, 500, 10))
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")

        ############################################
        self.widgetConnect()
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def widgetConnect(self):
        self.buttonGetContract.clicked.connect(self.getContract)
        self.buttonGetInput.clicked.connect(self.getInputFilesPath)
        self.buttonGetOutput.clicked.connect(self.getOutputFilesPath)
        self.radioButtonConsolid.toggled.connect(lambda: self.changeSecondMode())
        self.radioButtonDevice.toggled.connect(lambda: self.changeMode(mode='device'))
        self.buttonGenerate_1.clicked.connect(lambda: generateDeviceOutuput(mode='device'))
        self.buttonGenerate.clicked.connect(genarateContractOutput)


    def deviceCodeChanged(self, text, mode = 'device'):
        global curDeviceCode, curFile, contractFilesError, checkDeviceCode
        contractFilesError = False
        if text.split('-')[0] == 'ВШ':
            for file in devicesList:
                if f'ВП {text.split("-")[0]}-{text.split("-")[1]}' in file or f'ВП {text.split("-")[0]}-{text.split("-")[1]}' in file:
                    curDeviceCode = text
                    curFile = file
                    setVars(curDeviceCode)
                    break
                else:
                    curDeviceCode = ''

            if curDeviceCode == '':
                if mode == 'device':
                    self.showErrorMessagebox(text='Файл-шаблон\nотсутствует')
                    checkDeviceCode = False
                else:
                    contractFilesError = True
        else:
            for file in devicesList:
                if f'ВП {text}.xlsx' == file or f'ВП {text}.xls' == file:
                    curDeviceCode = text
                    curFile = file
                    checkDeviceCode = True
                    break
                else:
                    curDeviceCode = ''
            if curDeviceCode == '':
                if mode == 'device':
                    self.showErrorMessagebox(text='Файл-шаблон\nотсутствует')
                    checkDeviceCode = False
                else:
                    contractFilesError = True




    def changeMode(self, mode):
        global curMode
        global curSecondMode
        if self.radioButtonDevice.isChecked():
            curMode = 'Device'
            self.radioButtonConsolid.setChecked(True)
            self.radioButtonDelen.setChecked(False)
            self.radioButtonConsolid.setEnabled(False)
            self.radioButtonDelen.setEnabled(False)
            self.buttonGetContract.setStyleSheet("border: 1px solid;\n"
                                            "border-radius: 4px;\n"
                                            "border-color: #B7B7B7;\n"
                                            "background-color: #F4F4F4\n")
            self.comboBoxGetDevice.setStyleSheet("QComboBox{\n"
                                                 "background-color: #fff;\n"
                                                 "border: 1px solid;\n"
                                                 "border-radius: 4px;\n"
                                                 "\n"
                                                 "}\n"
                                                 "QComboBox::drop-down{\n"
                                                 "width: 34px;\n"
                                                 "height: 36px;\n"
                                                 "top: 0px;\n"
                                                 "border: 1px solid;\n"
                                                 "border-left: 2px solid;\n"
                                                 "border-radius: 3px;\n"
                                                 "background-color: #fff;\n"
                                                 "\n"
                                                 "}\n"
                                                 "\n"
                                                 "QComboBox::down-arrow{\n"
                                                 "image: url(:/img/Polygon 2.svg)\n"
                                                 "}\n"
                                                 "")
            self.lineEditAmount.setStyleSheet("padding-left: 2px;\n"
                                              "top: 0px;\n"
                                              "border: 2px solid;\n"
                                              "border-radius: 3px;\n"
                                              "font-weight: 400;\n"
                                              "border-color: #000;\n"
                                              "background-color: #fff")
            self.buttonGenerate_1.setStyleSheet("QPushButton{\n"
                                                "border: 2px solid;\n"
                                                "color: #000;\n"
                                                "border-color: #000;\n"
                                                "border-radius: 6px;\n"
                                                # "color: #848484;\n"
                                                "\n"
                                                "}\n"
                                                "\n"
                                                "")
            self.buttonGenerate.setStyleSheet("QPushButton{\n"
                                              "border: 2px solid;\n"
                                              "color: #000;\n"
                                              "border-color: #000;\n"
                                              "border-radius: 6px;\n"
                                              # "color: #848484;\n"
                                              "\n"
                                              "}\n"
                                              "\n"
                                              "")
            self.buttonGenerate.setStyleSheet("QPushButton{\n"
                                              "border: 2px solid;\n"
                                              "color: #B7B7B7;\n"
                                              "border-color: #B7B7B7;\n"
                                              "border-radius: 6px;\n"
                                              # "color: #848484;\n"
                                              "\n"
                                              "}\n"
                                              "\n"
                                              "")
            self.buttonGetContract.setEnabled(False)
            self.buttonGenerate.setEnabled(False)
            self.labelContract.setEnabled(False)
            self.comboBoxGetDevice.setEnabled(True)
            self.labelDeviceCode.setEnabled(True)
            self.lineEditAmount.setEnabled(True)
            self.buttonGenerate_1.setEnabled(True)
        else:
            curMode = 'Contract'
            curSecondMode = 'Consolid'
            self.radioButtonConsolid.setChecked(True)
            self.radioButtonDelen.setChecked(False)
            self.radioButtonConsolid.setEnabled(True)
            self.radioButtonDelen.setEnabled(True)
            self.comboBoxGetDevice.setStyleSheet("QComboBox{\n"
                                                 "background-color: #fff;\n"
                                                 "border: 1px solid;\n"
                                                 "border-radius: 4px;\n"
                                                 "border-color: #B7B7B7;\n"
                                                 "\n"
                                                 "}\n"
                                                 "QComboBox::drop-down{\n"
                                                 "width: 34px;\n"
                                                 "height: 36px;\n"
                                                 "top: 0px;\n"
                                                 "border: 1px solid;\n"
                                                 "border-left: 2px solid;\n"
                                                 "border-color: #B7B7B7;\n"
                                                 "border-radius: 3px;\n"
                                                 "background-color: #fff;\n"
                                                 "\n"
                                                 "}\n"
                                                 "\n"
                                                 "QComboBox::down-arrow{\n"
                                                 "image: url(:/img/PolygonGray.svg)\n"
                                                 "}\n"
                                                 "")
            self.buttonGetContract.setStyleSheet("border: 2px solid;\n"
                                                 "border-radius: 4px;\n"
                                                 "border-color: #858585;\n"
                                                 "background-color: #D9D9D9\n")
            self.buttonGenerate.setStyleSheet("QPushButton{\n"
                                              "border: 2px solid;\n"
                                              "color: #000;\n"
                                              "border-color: #000;\n"
                                              "border-radius: 6px;\n"
                                              #"color: #848484;\n"
                                              "\n"
                                              "}\n"
                                              "\n"
                                              "")
            self.buttonGenerate_1.setStyleSheet("QPushButton{\n"
                                              "border: 2px solid;\n"
                                              "color: #B7B7B7;\n"
                                              "border-color: #B7B7B7;\n"
                                              "border-radius: 6px;\n"
                                              # "color: #848484;\n"
                                              "\n"
                                              "}\n"
                                              "\n"
                                              "")
            self.lineEditAmount.setStyleSheet("padding-left: 2px;\n"
                                              "top: 0px;\n"
                                              "border: 2px solid;\n"
                                              "border-radius: 3px;\n"
                                              "font-weight: 400;\n"
                                              "border-color: #B7B7B7;\n"
                                              "background-color: #fff")
            self.buttonGenerate.setEnabled(True)
            self.buttonGetContract.setEnabled(True)
            self.labelContract.setEnabled(True)
            self.comboBoxGetDevice.setEnabled(False)
            self.labelDeviceCode.setEnabled(False)
            self.lineEditAmount.setEnabled(False)
            self.buttonGenerate_1.setEnabled(False)

    def changeSecondMode(self):
        global curSecondMode
        if self.radioButtonConsolid.isChecked():
            curSecondMode = 'Consolid'
        else:
            curSecondMode = 'Separate'
        print(curSecondMode)


    def getContract(self):
        global contractFilePath, curContractFile
        try:
            filePath = QtWidgets.QFileDialog.getOpenFileName()[0]
        except:
            pass
        if filePath!='':
            contractFilePath = filePath
            curContractFile = contractFilePath[contractFilePath.rfind('/') + 1:]


    def getInputFilesPath(self):
        global inputFilesPath
        global devicesList
        try:
            filesPath = QtWidgets.QFileDialog.getExistingDirectory()
            if filesPath!='':
                inputFilesPath=filesPath
                if self.checkFilesPath(inputFilesPath):
                    devicesList = os.listdir(inputFilesPath)
                    if 'Перечень изделий ЗАО ЗЭТ.txt' in devicesList:
                        devicesList.remove('Перечень изделий ЗАО ЗЭТ.txt')
                        self.pasteDevicesCodes(inputFilesPath)
                    else:
                        self.showErrorMessagebox(text='Не найден файл с\nкодами устройств')
                        self.comboBoxGetDevice.clear()
        except:
            pass


    def getOutputFilesPath(self):
        global outputFilesPath
        try:
            filesPath = QtWidgets.QFileDialog.getExistingDirectory()
            if filesPath!='':
                outputFilesPath = filesPath
                if self.checkFilesPath(outputFilesPath):
                    pass
                else:
                    outputFilesPath = ''
        except:
            pass

    def checkFilesPath(self, path):
        if os.access(path, mode=os.F_OK):
            if os.access(path, mode=os.X_OK):
                return True
            else:
                self.showErrorMessagebox(text='Нет доступа к директории')
                return False
        else:
            self.showErrorMessagebox(text='Такой директории не существует')
            return False

    def pasteDevicesCodes(self, path):
        with open(f"{path}/Перечень изделий ЗАО ЗЭТ.txt", encoding='utf-8') as file:
            for item in file:
                self.comboBoxGetDevice.addItem(item.replace('\n',''))
    def showErrorMessagebox(self, mode='default', text=''):
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Critical)
        msg.setText(text)
        msg.setWindowTitle("Information MessageBox")
        if mode == 'default':
            retval = msg.exec_()
        elif mode == 'connection':
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            text = 'Отсутствует\nинтернет-соединение'
            msg.setStandardButtons(QtWidgets.QMessageBox.Retry | QtWidgets.QMessageBox.Ok)
            buttonRetry = msg.button(QtWidgets.QMessageBox.Retry)
            msg.setText(text)
            buttonRetry.setText('Повторить')
            retval = msg.exec_()
            if retval == QtWidgets.QMessageBox.Retry:
                connect(mode='recheck')

    def showFinalMessage(self, text='Файл сгенерирован'):
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.setText(text)
        msg.setWindowTitle("Готово")
        retval = msg.exec_()


    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.buttonGenerate.setText(_translate("MainWindow", "Сгенерировать"))
        self.buttonGenerate_1.setText(_translate("MainWindow", "Сгенерировать"))
        self.labelGetInput.setText(_translate("MainWindow", "Выбор набора шаблонов"))
        self.buttonGetOutput.setText(_translate("MainWindow", "..."))
        self.buttonGetInput.setText(_translate("MainWindow", "..."))
        self.labelGetOutput.setText(_translate("MainWindow", "Сохранять в"))
        self.labelDeviceCode.setText(_translate("MainWindow", "Выбор изделия"))
        self.labelContract.setText(_translate("MainWindow", "Выбор договора"))
        self.buttonGetContract.setText(_translate("MainWindow", "..."))
        self.radioButtonDevice.setText(_translate("MainWindow", "Изделие"))
        self.radioButtonContract.setText(_translate("MainWindow", "Договор"))
        self.radioButtonConsolid.setText(_translate("MainWindow", "Консолидация"))
        self.radioButtonDelen.setText(_translate("MainWindow", "Деление"))



if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    connect()
    sys.exit(app.exec_())
